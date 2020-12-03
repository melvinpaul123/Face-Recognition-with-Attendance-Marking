import cv2   
import numpy as np
import face_recognition
import os
import sys
import sqlite3
import glob
from datetime import datetime
import openpyxl

path = os.path.dirname(os.path.abspath(__file__))

def facedatasetgenerator(cap):
    i=1
    def info(Id,Section,Name,Age,Gen):
        conn=sqlite3.connect(path+r'\SQLiteStudio\details.db')
        cmd="SELECT * FROM People WHERE ID="+str(Id)
        cursor=conn.execute(cmd)
        isRecordExist=0
        for row in cursor:
            isRecordExist=1
        if(isRecordExist==1):
            cmd="UPDATE People SET SECTION=' "+str(Section)+" 'WHERE ID='"+str(Id)+"' "
            cmd2="UPDATE People SET NAME=' "+str(Name)+" 'WHERE ID='"+str(Id)+"' "
            cmd3="UPDATE People SET AGE=' "+str(Age)+" 'WHERE ID='"+str(Id)+"' "
            cmd4="UPDATE People SET GENDER=' "+str(Gen)+" 'WHERE ID='"+str(Id)+"' "
        else:
            cmd="INSERT INTO People(ID,SECTION,NAME,AGE,GENDER) Values("+str(Id)+",'"+str(Section)+"','"+str(Name)+"','"+str(Age)+"','"+str(Gen)+"')"
            cmd2=""
            cmd3=""
            cmd4=""
        conn.execute(cmd)
        conn.execute(cmd2)
        conn.execute(cmd3)
        conn.execute(cmd4)
        conn.commit()
        conn.close()
    
    id=input('enter the Id= ')
    section=input('enter the Section= ').upper()
    name=input('enter the Name= ').title()
    age=input('enter the Age= ')
    gender=input('enter the Gender(M/F/Other)= ').upper()
    info(id,section,name,age,gender)
    
    if not os.path.exists('facedataset/'+str(section)+"/"+ str(id)):
        os.makedirs('facedataset/'+str(section)+"/"+str(id))
    
    cv2.namedWindow("img")
    cv2.resizeWindow("img",640,480)
    
    while 1:
        ret, img = cap.read()
        disp=img.copy()
        if ret == False:
            print('Failed to capture frame from camera. Check camera index in cv2.VideoCapture(0) \n')
            cv2.destroyAllWindows()
            break
        #img = cv2.resize(img, (0, 0), fx=0.25, fy=0.25)
        imgc=cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
        faceLoc = face_recognition.face_locations(imgc)
        if faceLoc ==[]:
            print("no face")
        else:
            #print(faceLoc[0])
            for faces in faceLoc:
                y1,x2,y2,x1=faces
                cv2.rectangle(disp,(x1, y1), (x2, y2),(255,0,255),2) # top, right, bottom, left
                face  = img[y1-10:y2+10, x1-10:x2+10]
        
        cv2.imshow('img',disp)
       
        k = cv2.waitKey(30) & 0xff
        if k == 27:
            cv2.destroyAllWindows()
            break
        if k == ord('p'): #pause stream
            print("-> Pausing Video Stream")
            print("-> Press any key to continue Video Stream")
            cv2.waitKey(-1) #wait until any key is pressed
        if k == ord('s'):
            cv2.imwrite("facedataset/"+str(section)+"/"+str(id)+"/"+str(id)+'.'+ str(i) + ".png", face)
            i=i+1
            
def facedetectoridentify(cap):
    section=input('enter the Section= ').upper()
    
    def getProfile(nbr_predicted):
        conn=sqlite3.connect(path+r'\SQLiteStudio\details.db')
        cmd="SELECT * FROM People WHERE ID="+str(nbr_predicted)
        cursor=conn.execute(cmd)
        profile=None
        for row in cursor:
            profile=row
        conn.close()
        return profile
    
    datasetpath = path+r"/facedataset/"+str(section)+"/"
    
    def get_images_and_labels(datasetpath):
     image_paths = [os.path.join(datasetpath, f) for f in os.listdir(datasetpath)]
     # images will contains face images
     images = []
     labels = []
     for image_path in image_paths:
         for image in glob.glob(image_path+r'/*.png'):
             img= cv2.imread(image)
             cv2.imshow("dataset",img)
             cv2.waitKey(100)
             # Get the label of the image
             nbr = int(os.path.split(image)[1].split(".")[0])
             print(nbr)
             labels.append(nbr)
             images.append(img)
         cv2.destroyAllWindows()
     return images, labels
    
    images, labels= get_images_and_labels(datasetpath)
    if images ==[]:
            print("no images in data set")
            sys.exit()
    
    def findEncodings(images):
        encodeList = []
        faceList = []
        for img in images:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            faceLoc = face_recognition.face_locations(img)
            if faceLoc ==[]:
                print("no other faces in the image dataset")
            else:
                encode = face_recognition.face_encodings(img)[0]
                encodeList.append(encode)
                faceList.append(faceLoc)
        return encodeList,faceList
    
    encodeListKnown,facesfound = findEncodings(images)
    if facesfound == []:
        print("no faces in the image dataset")
        sys.exit()
    print('Encodings Complete')
    
    def createxls(section):
        now = datetime.now()
        dt_string = now.strftime("%Y-%m-%d-%H")
        conn=sqlite3.connect(path+r'\SQLiteStudio\details.db')
        cur=conn.cursor()
        cur.execute("SELECT * FROM People")
        datas=cur.fetchall()
        headings=["ID","SECTION","NAME","AGE","GENDER","STATUS"]
        if os.path.exists(""'attendance/attendance for '+str(section)+'.xlsx'""):
            workbook = openpyxl.load_workbook(""'attendance/attendance for '+str(section)+'.xlsx'"")
            try:
                worksheet= workbook[dt_string]
            except KeyError:
                worksheet = workbook.create_sheet(dt_string,0)
            z=1
            for head in headings:
                worksheet.cell(row=1, column=z).value=head
                z+=1
            x=2
            y=1
            for row in datas:
                for info in row:
                    worksheet.cell(row=x, column=y).value=info
                    y+=1
                worksheet.cell(row=x, column=y).value="ABSENT"
                y=1
                x+=1
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.create_sheet(dt_string,0)
            z=1
            for head in headings:
                worksheet.cell(row=1, column=z).value=head
                z+=1
            x=2
            y=1
            for row in datas:
                for info in row:
                    worksheet.cell(row=x, column=y).value=info
                    y+=1
                worksheet.cell(row=x, column=y).value="ABSENT"
                y=1
                x+=1
        workbook.save(""'attendance/attendance for '+str(section)+'.xlsx'"")
        conn.close()
            
    createxls(section)
        
    def markAttendance(Id,section):
        now = datetime.now()
        dt_string = now.strftime("%Y-%m-%d-%H")
        workbook = openpyxl.load_workbook(""'attendance/attendance for '+str(section)+'.xlsx'"")
        sheet = workbook[dt_string]
        for x in range(2,sheet.max_row+1):
            if (sheet.cell(row=x, column=1).value==Id) and (sheet.cell(row=x, column=6).value=="ABSENT"):
                sheet.cell(row=x, column=6).value="PRESENT"
                workbook.save(""'attendance/attendance for '+str(section)+'.xlsx'"")
        
        #conn=sqlite3.connect(path+r'\SQLiteStudio\details.db')
        #cmd="SELECT * FROM People WHERE ID="+str(Id)+""
        #cursor=conn.execute(cmd)
        #for row in cursor:
         #   if(row[4]!=0):
          #      cmd="UPDATE People SET Attendance=' "+str(dt_string)+" 'WHERE ID='"+str(Id)+"'"
           # else:
            #    cmd="INSERT INTO People SET Attendance=' "+str(dt_string)+" 'WHERE ID='"+str(Id)+"'"
        #conn.execute(cmd)
        #conn.commit()
        #conn.close()
        
            
    while 1:
        ret, imgq = cap.read()
        if ret == False:
            print('Failed to capture frame from camera. Check camera index in cv2.VideoCapture(0) \n')
            cv2.destroyAllWindows()
            break
        img=imgq.copy()
        img = cv2.resize(img, (0, 0), fx=0.25, fy=0.25)
        imgw=cv2.cvtColor(img,cv2.COLOR_BGR2RGB)
        faceLocw = face_recognition.face_locations(imgw)
        encodew = face_recognition.face_encodings(imgw, faceLocw)
        
        for encodeFace,faceLoc in zip(encodew,faceLocw):
            #matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)
            if faceDis[matchIndex]< 0.40:
                #print(faceDis[matchIndex])
                id = labels[matchIndex]
                profile=getProfile(id)
                markAttendance(id,section)
                print(profile[2])
                y1,x2,y2,x1=faceLoc
                y1, x2, y2, x1 = y1*4,x2*4,y2*4,x1*4
                cv2.rectangle(imgq, (x1, y1), (x2, y2), (0, 255, 0), 2)
                #cv2.rectangle(imgq, (x1, y2 - 35), (x2, y2), (0, 255, 0),-1)
                cv2.putText(imgq, "Name:"+str(profile[2]), (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 255, 255), 1)
                #cv2.putText(imgq, "Id:"+str(profile[0]), (x1 + 6, y2+20), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 255, 255), 1)
                #cv2.putText(imgq, "Section:"+str(profile[1]), (x1 + 6, y2+20), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 255, 255), 1)
                #cv2.putText(img, "Age:"+str(profile[3]), (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 255, 255), 1)
                #cv2.putText(img, "Gender:"+str(profile[4]), (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 255, 255), 1)
            else:
                cv2.rectangle(imgq, (x1, y1), (x2, y2), (0, 0, 255), 2)
                #cv2.rectangle(imgq, (x1, y2 - 35), (x2, y2), (0, 0, 255), cv2.FILLED)
                cv2.putText(imgq, "Unknown", (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 255, 255), 1)
        cv2.imshow("img",imgq)
        k = cv2.waitKey(10) & 0xff
        if k == 27:
            cv2.destroyAllWindows()
            break
        if k == ord('p'): #pause stream
            print("-> Pausing Video Stream")
            print("-> Press any key to continue Video Stream")
            cv2.waitKey(-1) #wait until any key is pressed


#capturing video through webcam
print("-> Starting Video Stream...")
print("-> Press Q to Exit the Program")
cap=cv2.VideoCapture('obama.mp4')

# We need to check if camera is opened previously or not 
if (cap.isOpened() == False):  
    print("-> Error reading video file") 
    sys.exit()

# We need to set resolutions. 
# so, convert them from float to integer. 
frame_width = int(cap.get(3)) 
frame_height = int(cap.get(4)) 
   
size = (frame_width, frame_height) 
   
# Below VideoWriter object will create 
# a frame of above defined The output  
# is stored in 'filename.avi' file. 
#fourcc = cv2.VideoWriter_fourcc(*'XVID')
#result = cv2.VideoWriter('result.avi', fourcc, 25.0, size)


print("1: ADD NEW FACE TO DATASET\n")
print("2: IDENTIFY THE PERSON\n")
per="q"
while per != "n":
    z= input("SELECT THE OPTION: ")
    if(z=="1"):
        facedatasetgenerator(cap)
    elif(z=="2"):
        facedetectoridentify(cap)
    else:
        print("WRONG OPTION")
    per= input("DO YOU WANT TO CONTINUE -> y/n: ")

print("-> Ending Video Stream")
cap.release()